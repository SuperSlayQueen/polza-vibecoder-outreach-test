#!/usr/bin/env python3
"""Финальная чистая сдача без демо-листов."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def add_df(wb: Workbook, title: str, df: pd.DataFrame, idx: int | None = None) -> None:
    ws = wb.create_sheet(title, idx if idx is not None else len(wb.sheetnames))
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for col in ws.columns:
        letter = col[0].column_letter
        width = min(55, max(12, max(len(str(cell.value or "")) for cell in col[:30]) + 2))
        ws.column_dimensions[letter].width = width


def main() -> None:
    base_raw = pd.read_excel("data/polza_outreach_base.xlsx")
    hard = pd.read_csv("data/companies_hardened.csv")
    hard["website_norm"] = hard["website"].astype(str).str.lower().str.rstrip("/")
    base_raw["website_norm"] = base_raw["Сайт"].astype(str).str.lower().str.rstrip("/")
    base = base_raw.merge(
        hard[["website_norm", "mx_ok", "mx", "russia_focus"]],
        on="website_norm",
        how="left",
    )

    base_out = pd.DataFrame(
        {
            "Компания": base["Компания"],
            "Сайт": base["Сайт"],
            "Имя": base["Имя"],
            "Email": base["Email"],
            "Должность": base["Должность"],
            "Ниша": base["Ниша"],
            "Персонализация": base["Персонализация"],
            "Тип_контакта": "role-based (публичный корпоративный ящик)",
            "MX_домена": base["mx"].fillna(""),
            "MX_ok": base["mx_ok"].fillna(False),
            "РФ_домен": base["russia_focus"].fillna(False),
            "Источник_персонализации": base["Источник_персонализации"],
            "QA_флаги": base["QA_флаги"].fillna(""),
        }
    )
    # убрать строки без персонализации
    base_out = base_out[base_out["Персонализация"].astype(str).str.len() > 10]
    base_out = base_out.drop_duplicates(subset=["Сайт"])
    assert len(base_out) >= 50, len(base_out)

    letters = pd.DataFrame(
        [
            {
                "Письмо": 1,
                "Когда": "День 0",
                "Тема": "Входящие от ЛПР без увеличения рекламного бюджета",
                "Тело": (
                    "{{имя}}, добрый день.\n\n{{персонализация}}\n\n"
                    "Мы в Polza Agency собираем базы ЛПР и запускаем персональные email-цепочки "
                    "для B2B-компаний — чтобы заявки шли не только из контекста и рекомендаций.\n\n"
                    "Обычно первые предметные диалоги появляются за 10–14 дней после старта. "
                    "Если у вас сейчас упираетесь в поток квалифицированных лидов — могу коротко "
                    "показать, как это выглядит на похожих нишах.\n\n"
                    "Удобно 15 минут на этой неделе?"
                ),
            },
            {
                "Письмо": 2,
                "Когда": "+3 дня",
                "Тема": "Re: входящие от ЛПР — цифры по похожим запускам",
                "Тело": (
                    "{{имя}}, ещё раз коротко — без давления.\n\n"
                    "Не «база на продажу», а цикл: гипотеза → ЛПР → персонализация → цепочка → "
                    "квалификация ответов. На запусках вроде StaffLine / digital и B2B-опта "
                    "открываемость часто 70–90%, ответы 4–8%, дальше — уже предметный интерес "
                    "(КП, демо, расчёт).\n\n"
                    "Если канал для вас не приоритет — ок, просто скажите. Если приоритет есть, "
                    "но руки не доходят собрать это внутри — могу прислать 1-страничный разбор "
                    "под ваш сегмент."
                ),
            },
            {
                "Письмо": 3,
                "Когда": "+5 дней после письма 2",
                "Тема": "Закрываю вопрос по аутричу",
                "Тело": (
                    "{{имя}}, последнее письмо по теме — не хочу занимать место в почте зря.\n\n"
                    "Если холодный аутрич до ЛПР вам сейчас не нужен или закрываете лиды другими "
                    "каналами — просто проигнорируйте, больше не напишу.\n\n"
                    "Если отложили «на потом»: напишите «интересно» или «позже» — вернусь в "
                    "удобный момент с коротким расчётом воронки под ваш средний чек и объём базы.\n\n"
                    "Удачных запусков."
                ),
            },
        ]
    )

    t4 = pd.read_excel("data/task4_final.xlsx")

    wb = Workbook()
    wb.remove(wb.active)
    add_df(wb, "База", base_out, 0)
    add_df(wb, "Task4", t4, 1)
    add_df(wb, "Письма", letters, 2)

    ws = wb.create_sheet("Цепочка (текст)", 3)
    ws["A1"] = Path("email_sequence.md").read_text(encoding="utf-8")
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 100

    ws = wb.create_sheet("Методология", 4)
    ws["A1"] = Path("METHODOLOGY.md").read_text(encoding="utf-8")
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 100

    ws = wb.create_sheet("Task4_разбор", 5)
    ws["A1"] = Path("TASK4_NOTES.md").read_text(encoding="utf-8")
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 100

    # подсветка ловушек Task4
    red = PatternFill("solid", fgColor="FFE5E5")
    yellow = PatternFill("solid", fgColor="FFF6D5")
    green = PatternFill("solid", fgColor="E5F7E8")
    ws_t4 = wb["Task4"]
    status_idx = None
    for i, cell in enumerate(ws_t4[1], 1):
        if cell.value == "Статус":
            status_idx = i
    if status_idx:
        for row in ws_t4.iter_rows(min_row=2, max_row=ws_t4.max_row):
            val = str(row[status_idx - 1].value or "")
            fill = red if "ЛОВУШКА" in val else yellow if ("СОМНЕНИЕ" in val or "оговорк" in val) else green if val.startswith("OK") else None
            if fill:
                for c in row:
                    c.fill = fill

    dest = Path("data/Polza_Test_Submission.xlsx")
    wb.save(dest)
    base_out.to_csv("data/polza_base_final.csv", index=False, encoding="utf-8-sig")
    print(f"rows={len(base_out)} mx_ok={int(base_out['MX_ok'].sum())} -> {dest}")
    print("sheets:", wb.sheetnames)


if __name__ == "__main__":
    main()
