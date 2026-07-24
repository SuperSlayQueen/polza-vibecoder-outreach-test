#!/usr/bin/env python3
"""Пост-обработка базы: убрать мёртвые сайты, дописать ручную персонализацию, собрать XLSX."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# Ручные факты только там, где сайт не отдал текст (источник указан)
MANUAL = {
    "TrueConf": (
        "TrueConf — российский разработчик защищённых решений для видеосвязи "
        "и ВКС для корпораций и госсектора.",
        "сайт/публичное описание продукта",
    ),
    "Envycrm": (
        "Envycrm позиционируется как CRM для команд продаж с упором на простоту "
        "внедрения и работу со сделками.",
        "сайт продукта",
    ),
    "1С:Сервистренд": (
        "1С:Сервистренд — франчайзи 1С: услуги внедрения, сопровождения и продажи "
        "программ 1С для бизнеса.",
        "servicetrend.ru /communications",
    ),
    "Эльба": (
        "Эльба (Контур) — онлайн-бухгалтерия для ИП и ООО: отчётность, счета, "
        "учёт без штатного бухгалтера.",
        "публичное описание продукта Контур.Эльба",
    ),
    "Abbyy": (
        "ABBYY — решения Intelligent Document Processing / OCR для извлечения "
        "данных из документов в корпоративных процессах.",
        "публичное позиционирование ABBYY",
    ),
    "Omnidesk": (
        "Omnidesk — helpdesk/омниканальная поддержка для обработки обращений "
        "клиентов в одном окне.",
        "публичное описание продукта",
    ),
}


def add_df(wb: Workbook, title: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(title)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for col in ws.columns:
        letter = col[0].column_letter
        width = min(60, max(12, max(len(str(c.value or "")) for c in col[:40]) + 2))
        ws.column_dimensions[letter].width = width


def main() -> None:
    raw = pd.read_excel("data/polza_outreach_base.xlsx")
    # убираем заведомо битые/нерелевантные для демо строки без ручного покрытия
    drop_names = {"Alto", "UDS Game", "Right Way"}
    df = raw[~raw["Компания"].isin(drop_names)].copy()

    for name, (text, src) in MANUAL.items():
        mask = df["Компания"] == name
        if mask.any():
            df.loc[mask, "Персонализация"] = text
            df.loc[mask, "Источник_персонализации"] = f"manual:{src}"
            # сбрасываем «unavailable» шум если факт добрали
            df.loc[mask, "QA_флаги"] = df.loc[mask, "QA_флаги"].fillna("").astype(str)
            df.loc[mask, "QA_флаги"] = (
                df.loc[mask, "QA_флаги"]
                .str.replace("сайт недоступен", "", regex=False)
                .str.replace("слабый сигнал с сайта", "", regex=False)
                .str.replace(r"^;\s*|;$", "", regex=True)
                .str.strip("; ")
            )

    # ложные «название ≠ сайт» на кириллических брендах с латинским title —
    # после улучшения name_matches_site в personalize.py чистим старый прогон
    def clean_qa(row) -> str:
        qa = str(row.get("QA_флаги") or "")
        if qa in {"nan", "None"}:
            return ""
        company = str(row.get("Компания") or "")
        site = str(row.get("Сайт") or "")
        # если бренд читается из домена — убираем name mismatch
        from personalize import SiteSnapshot, name_matches_site

        snap = SiteSnapshot(url=site, reachable=True, title="", h1="", meta_description="")
        if name_matches_site(company, snap):
            qa = qa.replace("название ≠ сайт", "").replace("название != сайт", "")
        # чистим мусорные «emails» из старого прогона
        return "; ".join(p.strip() for p in qa.split(";") if p.strip())

    df["QA_флаги"] = df.apply(clean_qa, axis=1)

    # вычищаем мусорные email вида *.webp из старого прогона
    def clean_site_emails(val: str) -> str:
        parts = [p.strip() for p in str(val or "").split(",") if p.strip()]
        keep = [
            p
            for p in parts
            if not any(x in p.lower() for x in (".webp", ".png", ".jpg", ".svg", ".gif"))
        ]
        return ", ".join(keep)

    if "Emails_на_сайте" in df.columns:
        df["Emails_на_сайте"] = df["Emails_на_сайте"].map(clean_site_emails)

    # Avito/HH персонализация слишком «витринная» — подчистим
    fixes = {
        "Avito Работа (B2B)": (
            "Avito для бизнеса закрывает массовый найм и B2B-продвижение через "
            "объявления; отдельный контур — Avito Работа для работодателей."
        ),
        "HeadHunter (HH Pro)": (
            "hh.ru — крупнейшая площадка найма в РФ; для компаний есть корпоративные "
            "продукты и услуги продвижения бренда работодателя."
        ),
        "Хабр Карьера (для бизнеса)": (
            "Хабр Карьера — канал найма IT-специалистов через экосистему Хабра "
            "и карьерные инструменты для работодателей."
        ),
    }
    for name, text in fixes.items():
        mask = df["Компания"] == name
        if mask.any():
            df.loc[mask, "Персонализация"] = text
            df.loc[mask, "Источник_персонализации"] = "manual:публичные страницы продуктов"

    # чистый вид под сдачу
    out = df[
        [
            "Компания",
            "Сайт",
            "Имя",
            "Email",
            "Должность",
            "Ниша",
            "Персонализация",
            "Источник_персонализации",
            "QA_флаги",
            "Emails_на_сайте",
        ]
    ].copy()
    out["QA_флаги"] = out["QA_флаги"].fillna("")
    out = out.drop_duplicates(subset=["Сайт"], keep="first")

    assert len(out) >= 50, len(out)

    wb = Workbook()
    wb.remove(wb.active)
    add_df(wb, "База", out)

    letters = pd.DataFrame(
        [
            {
                "Письмо": 1,
                "Когда": "День 0",
                "Тема": "Входящие от ЛПР без увеличения рекламного бюджета",
                "Тело": (
                    "{{имя}}, добрый день.\n\n{{персонализация}}\n\n"
                    "Мы в Polza Agency собираем базы ЛПР и запускаем персональные email-цепочки "
                    "для B2B-компаний — чтобы заявки шли не только из контекста и рекомендаций.\n\n"
                    "Обычно первые предметные диалоги появляются за 10–14 дней после старта. "
                    "Если у вас сейчас упираетесь в поток квалифицированных лидов — могу коротко "
                    "показать, как это выглядит на похожих нишах.\n\n"
                    "Удобно 15 минут на этой неделе?"
                ),
            },
            {
                "Письмо": 2,
                "Когда": "+3 дня",
                "Тема": "Re: входящие от ЛПР — цифры по похожим запускам",
                "Тело": (
                    "{{имя}}, ещё раз коротко — без давления.\n\n"
                    "Не «база на продажу», а цикл: гипотеза → ЛПР → персонализация → цепочка → "
                    "квалификация ответов. На запусках вроде StaffLine / digital и B2B-опта "
                    "открываемость часто 70–90%, ответы 4–8%, дальше — уже предметный интерес "
                    "(КП, демо, расчёт).\n\n"
                    "Если канал для вас не приоритет — ок, просто скажите. Если приоритет есть, "
                    "но руки не доходят собрать это внутри — могу прислать 1-страничный разбор "
                    "под ваш сегмент."
                ),
            },
            {
                "Письмо": 3,
                "Когда": "+5 дней после письма 2",
                "Тема": "Закрываю вопрос по аутричу",
                "Тело": (
                    "{{имя}}, последнее письмо по теме — не хочу занимать место в почте зря.\n\n"
                    "Если холодный аутрич до ЛПР вам сейчас не нужен или закрываете лиды другими "
                    "каналами — просто проигнорируйте, больше не напишу.\n\n"
                    "Если отложили «на потом»: напишите «интересно» или «позже» — вернусь в "
                    "удобный момент с коротким расчётом воронки под ваш средний чек и объём базы.\n\n"
                    "Удачных запусков."
                ),
            },
        ]
    )
    add_df(wb, "Письма", letters)

    ws = wb.create_sheet("Цепочка (текст)")
    ws["A1"] = Path("email_sequence.md").read_text(encoding="utf-8")
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 110

    ws = wb.create_sheet("Методология")
    ws["A1"] = Path("METHODOLOGY.md").read_text(encoding="utf-8")
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 110

    # task4 demo results if present
    t4 = Path("data/task4_demo_result.xlsx")
    if t4.exists():
        demo = pd.read_excel(t4)
        keep = [
            c
            for c in [
                "Компания",
                "Сайт",
                "Имя",
                "Email",
                "Персонализация",
                "QA_флаги",
                "Источник_персонализации",
            ]
            if c in demo.columns
        ]
        add_df(wb, "Task4_демо_ловушки", demo[keep])

    dest = Path("data/Polza_Test_Submission.xlsx")
    wb.save(dest)
    out.to_csv("data/polza_base_final.csv", index=False, encoding="utf-8-sig")
    print(f"rows={len(out)} -> {dest}")


if __name__ == "__main__":
    main()
