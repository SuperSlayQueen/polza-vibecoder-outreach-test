#!/usr/bin/env python3
"""Собрать лист Task4 с персонализацией + человеческим разбором ловушек."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Вердикт после сверки сайт↔название↔email (ручной ресёрч + флаги скрипта)
VERDICTS = {
    "Mingwen Intelligent": {
        "status": "OK (с оговоркой)",
        "trap": "Нет критичной путаницы: email и сайт на mgwmachine.com. Title на китайском (铭文机床) — скрипт помечает «название ≠ сайт», это ложное срабатывание из-за языка.",
        "personalization_fix": (
            "Mingwen Intelligent (浙江铭文智能科技) — китайский производитель станков; "
            "на mgwmachine.com позиционируется как производитель CNC/металлообрабатывающего оборудования."
        ),
    },
    "HNC": {
        "status": "OK",
        "trap": "Совпадает: hnc.su + sales@hnc.su. Сайт — HNC Electric, АСУТП/ЧПУ/привода.",
        "personalization_fix": "",
    },
    "Tengzhong Machinery": {
        "status": "ЛОВУШКА",
        "trap": (
            "Сайт uncomtech.ru принадлежит холдингу Ункомтех (кабель/провод), а не Tengzhong. "
            "Email sales01@nttzmt.com — от Nantong Tengzhong (nttzmt.com). "
            "Строки Tengzhong и ТД Ункомтех перепутаны местами по сайту."
        ),
        "personalization_fix": (
            "Nantong Tengzhong Machinery — китайский производитель гильотин, листогибов и вальцов "
            "(бренд Tengzhong); корпоративный домен nttzmt.com / tzmachinery.com."
        ),
        "fix_website": "https://en.nttzmt.com",
        "fix_email": "sales01@nttzmt.com",
    },
    "Tesid Equipment": {
        "status": "ЛОВУШКА / сомнение",
        "trap": (
            "Сайт unimach.ru открывается как «Юнимаш» (лазерный раскрой), не Tesid Equipment. "
            "Email sales@unimach.ru бьётся с сайтом, но название компании в строке другое — "
            "похоже на подмену имени или чужой бренд в поле company."
        ),
        "personalization_fix": (
            "По указанному сайту unimach.ru — компания «Юнимаш», производит установки "
            "лазерного раскроя металла. Связь с брендом Tesid по открытым данным не подтверждена."
        ),
    },
    "ТД Ункомтех": {
        "status": "ЛОВУШКА",
        "trap": (
            "Сайт saintymachine.com и email sales@thebestcnc.com не относятся к Ункомтех. "
            "Реальный сайт Ункомтех — uncomtech.ru (сейчас стоит у строки Tengzhong). "
            "saintymachine.com в таблице ещё раз у Shixinghong — дубль чужого сайта."
        ),
        "personalization_fix": (
            "Ункомтех — российский холдинг, крупный поставщик кабельно-проводниковой продукции "
            "(Иркутсккабель, Кирскабель); сайт uncomtech.ru."
        ),
        "fix_website": "https://uncomtech.ru",
        "fix_email": "sales@uncomtech.com",
    },
    "JAT Cemented Carbide": {
        "status": "ЛОВУШКА",
        "trap": (
            "Сайт/email jillionsupply.com — поставщик motors/actuators/servo, не твердый сплав JAT. "
            "Контакты JAT (jatcarbide.com / sales@jatcarbide.com) стоят у строки Rogen Technologies."
        ),
        "personalization_fix": (
            "Zhuzhou JAT Cemented Carbide — производитель твердосплавного CNC-инструмента "
            "(концевые фрезы, пластины); сайт jatcarbide.com / jat-carbide.com."
        ),
        "fix_website": "https://www.jatcarbide.com",
        "fix_email": "sales@jatcarbide.com",
    },
    "Rogen Technologies": {
        "status": "ЛОВУШКА",
        "trap": (
            "В строке стоят сайт/email JAT Carbide (jat-carbide.com, sales@jatcarbide.com). "
            "К Rogen Technologies эти контакты не относятся — классическая перестановка строк."
        ),
        "personalization_fix": (
            "По полям website/email в таблице фактически лежит JAT Carbide, не Rogen. "
            "Персонализацию по Rogen без отдельного источника не пишу — будет выдумка."
        ),
    },
    "Искролайн": {
        "status": "OK",
        "trap": "Совпадает. Российский производитель атомно-эмиссионных спектрометров.",
        "personalization_fix": "",
    },
    "Jimmy CNC Tool": {
        "status": "OK",
        "trap": "Совпадает: jimmytool.com, китайский производитель carbide end mills / CNC tools.",
        "personalization_fix": "",
    },
    "Ezhong Heavy Machinery": {
        "status": "OK",
        "trap": "Совпадает: ezhonggroup.com — металлообработка, press/rolling/leveling machines.",
        "personalization_fix": "",
    },
    "Shixinghong Precision": {
        "status": "ЛОВУШКА",
        "trap": (
            "Email swyct@126.com — китайский freemail (126.com), не корпоративный домен. "
            "Сайт saintymachine.com тот же чужой домен, что и у «ТД Ункомтех» — дубль. "
            "Строку нельзя использовать в рассылке без перепроверки."
        ),
        "personalization_fix": (
            "Достоверную персонализацию не ставлю: сайт недоступен/чужой, email на 126.com."
        ),
    },
    "Internor Machinery": {
        "status": "СОМНЕНИЕ",
        "trap": (
            "Email internor.com.cn vs сайт internor-mach.com — близкие, но разные домены "
            "(часто бывает у CN-экспортёров). Сайт в прогоне отдал SSLError — перепроверить вручную."
        ),
        "personalization_fix": (
            "Internor Machinery — китайский поставщик станков/оборудования (по названию и доменам "
            "internor-mach.com / internor.com.cn); сайт на момент сбора по SSL не открылся стабильно."
        ),
    },
    "РИЦ Техносфера": {
        "status": "СОМНЕНИЕ",
        "trap": (
            "Домен technosphera.ru и email совпадают, но выдача скрипта с главной слабая "
            "(мало продуктового сигнала / возможна витрина издательства «Техносфера»). "
            "Перед запуском — руками подтвердить, что это нужный B2B-поставщик станков."
        ),
        "personalization_fix": "",
    },
    "Howfit Science": {
        "status": "OK",
        "trap": "Совпадает: howfit-press.com — high-speed / stamping presses с 2006 г.",
        "personalization_fix": "",
    },
    "Fengyi Yinhu": {
        "status": "OK",
        "trap": "Совпадает: fengyitool.com — Chengdu Fengyi, резьбовой твердосплавный инструмент.",
        "personalization_fix": "",
    },
}


def main() -> None:
    src = pd.read_excel("data/task4_personalized.xlsx")
    rows = []
    for _, r in src.iterrows():
        company = str(r["Компания"])
        meta = VERDICTS.get(company, {})
        pers = str(r["Персонализация"] or "")
        fix = meta.get("personalization_fix") or ""
        if fix:
            pers = fix
        rows.append(
            {
                "Компания": company,
                "Сайт_как_в_задании": r["Сайт"],
                "Email_как_в_задании": r["Email"],
                "Имя": r["Имя"],
                "Персонализация": pers,
                "Статус": meta.get("status", ""),
                "Разбор_ловушки": meta.get("trap", ""),
                "Сайт_исправленный": meta.get("fix_website", ""),
                "Email_исправленный": meta.get("fix_email", ""),
                "QA_скрипта": r.get("QA_флаги", ""),
                "Title_с_сайта": r.get("Title", ""),
                "Источник": "website+manual QA" if fix else "website script",
            }
        )

    out = pd.DataFrame(rows)
    out_path = Path("data/task4_final.xlsx")
    out.to_excel(out_path, index=False)

    # вшить в общий submission
    submission = Path("data/Polza_Test_Submission.xlsx")
    if submission.exists():
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows as d2r
        import openpyxl

        wb = openpyxl.load_workbook(submission)
        for name in ("Task4", "Task4_разбор"):
            if name in wb.sheetnames:
                del wb[name]
        ws = wb.create_sheet("Task4", 1)
        for row in d2r(out, index=False, header=True):
            ws.append(row)
        for c in ws[1]:
            c.font = Font(bold=True)
        red = PatternFill("solid", fgColor="FFE5E5")
        yellow = PatternFill("solid", fgColor="FFF6D5")
        green = PatternFill("solid", fgColor="E5F7E8")
        status_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == "Статус":
                status_col = idx
        if status_col:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                val = str(row[status_col - 1].value or "")
                fill = None
                if "ЛОВУШКА" in val:
                    fill = red
                elif "СОМНЕНИЕ" in val or "оговорк" in val:
                    fill = yellow
                elif val.startswith("OK"):
                    fill = green
                if fill:
                    for cell in row:
                        cell.fill = fill
        ws.freeze_panes = "A2"

        notes = Path("TASK4_NOTES.md")
        analysis = """# Задача 4 — разбор

В таблице с 15 компаниями данные частично перепутаны между строками
(классическая ловушка «вдруг мы что-то перепутали»).

## Явные перестановки

1. **Tengzhong ↔ Ункомтех**: у Tengzhong стоит сайт uncomtech.ru (кабель),
   у Ункомтех — saintymachine.com + thebestcnc email.
2. **JAT ↔ Rogen**: у JAT стоят контакты jillionsupply, у Rogen — jatcarbide.
3. **Shixinghong**: freemail @126.com + тот же saintymachine.com (дубль).
4. **Tesid / Юнимаш**: сайт unimach.ru описывает Юнимаш, не Tesid.

## Что сдаём

Лист Task4 в Polza_Test_Submission.xlsx: персонализация + статус + разбор.
Строки с ловушками не выкидывал молча — пометил и предложил исправления где возможно.
"""
        notes.write_text(analysis, encoding="utf-8")
        ws2 = wb.create_sheet("Task4_разбор")
        ws2["A1"] = analysis
        ws2["A1"].alignment = Alignment(wrap_text=True, vertical="top")
        ws2.column_dimensions["A"].width = 110
        wb.save(submission)
        print(f"updated {submission}")

    print(out[["Компания", "Статус"]].to_string(index=False))
    print(f"saved {out_path}")


if __name__ == "__main__":
    main()
